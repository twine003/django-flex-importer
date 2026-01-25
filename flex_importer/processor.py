"""
Import processor for FlexImporter
"""
import csv
import json
from datetime import datetime, date, time
from decimal import Decimal
from io import TextIOWrapper, BytesIO
from openpyxl import load_workbook
from django.utils import timezone
from .models import ImportJob


def make_json_serializable(data):
    """
    Convert data to JSON-serializable format.
    Handles datetime, date, time, Decimal, and other non-serializable types.
    """
    if data is None:
        return None
    elif isinstance(data, dict):
        return {k: make_json_serializable(v) for k, v in data.items()}
    elif isinstance(data, (list, tuple)):
        return [make_json_serializable(item) for item in data]
    elif isinstance(data, datetime):
        return data.isoformat()
    elif isinstance(data, date):
        return data.isoformat()
    elif isinstance(data, time):
        return data.isoformat()
    elif isinstance(data, Decimal):
        return float(data)
    elif isinstance(data, bytes):
        return data.decode('utf-8', errors='replace')
    else:
        # Try to return as-is, but catch any serialization issues
        try:
            json.dumps(data)
            return data
        except (TypeError, ValueError):
            return str(data)


class ImportProcessor:
    """Process imports from different file formats"""

    def __init__(self, import_job):
        self.import_job = import_job
        self.importer_class = None

    def process(self):
        """Main process method to handle import"""
        try:
            from .registry import importer_registry

            self.importer_class = importer_registry.get_importer(
                self.import_job.importer_class
            )

            if not self.importer_class:
                self.import_job.status = 'failed'
                self.import_job.result_message = 'Clase importadora no encontrada'
                self.import_job.save()
                return False

            self.import_job.status = 'processing'
            self.import_job.started_at = timezone.now()
            self.import_job.add_progress_log('Iniciando importación...')
            self.import_job.save()

            file_format = self.import_job.file_format
            file_path = self.import_job.uploaded_file.path

            if file_format == 'xlsx':
                rows = self._read_xlsx(file_path)
            elif file_format == 'csv':
                rows = self._read_csv(file_path)
            elif file_format == 'json':
                rows = self._read_json(file_path)
            else:
                raise ValueError(f'Formato no soportado: {file_format}')

            self.import_job.total_rows = len(rows)
            self.import_job.add_progress_log(f'Se encontraron {len(rows)} filas para procesar')
            self.import_job.save()

            self._process_rows(rows)

            self.import_job.completed_at = timezone.now()

            if self.import_job.error_rows == 0:
                self.import_job.status = 'success'
                message_parts = [f'Importación completada exitosamente. {self.import_job.success_rows} filas procesadas']
                if self.import_job.updated_rows > 0 or self.import_job.created_rows > 0:
                    message_parts.append(f'({self.import_job.created_rows} creadas, {self.import_job.updated_rows} actualizadas)')
                self.import_job.result_message = ' '.join(message_parts) + '.'
            elif self.import_job.success_rows > 0:
                self.import_job.status = 'partial'
                message_parts = [f'Importación parcial. {self.import_job.success_rows} exitosas']
                if self.import_job.updated_rows > 0 or self.import_job.created_rows > 0:
                    message_parts.append(f'({self.import_job.created_rows} creadas, {self.import_job.updated_rows} actualizadas)')
                message_parts.append(f'{self.import_job.error_rows} con errores')
                self.import_job.result_message = ', '.join(message_parts) + '.'
            else:
                self.import_job.status = 'failed'
                self.import_job.result_message = f'Importación fallida. Todas las filas tuvieron errores.'

            self.import_job.add_progress_log(self.import_job.result_message, 'success' if self.import_job.status == 'success' else 'warning')
            self.import_job.save()

            return True

        except Exception as e:
            self.import_job.status = 'failed'
            self.import_job.result_message = f'Error en importación: {str(e)}'
            self.import_job.completed_at = timezone.now()
            self.import_job.add_progress_log(f'Error: {str(e)}', 'error')
            self.import_job.save()
            return False

    def _read_xlsx(self, file_path):
        """Read data from Excel file"""
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        # Obtener header_row desde la Meta del importador (default: 1)
        header_row = 1
        if hasattr(self.importer_class, 'Meta') and hasattr(self.importer_class.Meta, 'header_row'):
            header_row = self.importer_class.Meta.header_row

        headers = []
        for cell in ws[header_row]:
            if cell.value:
                header = str(cell.value).replace(' *', '').strip()
                headers.append(header)

        rows = []
        data_start_row = header_row + 1
        for row_idx, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
            if all(cell is None or cell == '' for cell in row):
                continue

            row_data = {}
            for idx, value in enumerate(row):
                if idx < len(headers):
                    row_data[headers[idx]] = value

            row_data['_row_number'] = row_idx
            rows.append(row_data)

        return rows

    def _read_csv(self, file_path):
        """Read data from CSV file"""
        rows = []

        with open(file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)

            headers = [h.replace(' *', '').strip() for h in reader.fieldnames]
            reader.fieldnames = headers

            for row_idx, row in enumerate(reader, start=2):
                if all(not value or value == '' for value in row.values()):
                    continue

                row['_row_number'] = row_idx
                rows.append(row)

        return rows

    def _read_json(self, file_path):
        """Read data from JSON file"""
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        if isinstance(data, dict) and 'data' in data:
            rows = data['data']
        elif isinstance(data, list):
            rows = data
        else:
            raise ValueError('Formato JSON inválido. Debe ser una lista o un objeto con propiedad "data"')

        for idx, row in enumerate(rows, start=1):
            row['_row_number'] = idx

        return rows

    def _process_rows(self, rows):
        """Process each row of data"""
        field_info = self.importer_class.get_field_info()
        field_name_map = {info['verbose_name']: info['name'] for info in field_info}

        importer_instance = self.importer_class()

        for idx, row_data in enumerate(rows, start=1):
            row_number = row_data.get('_row_number', idx)

            normalized_data = {}
            for key, value in row_data.items():
                if key == '_row_number':
                    continue

                field_name = field_name_map.get(key, key)
                normalized_data[field_name] = value

            validated_data, errors = importer_instance.validate_row(normalized_data)

            if errors:
                self.import_job.error_rows += 1
                error_entry = {
                    'row': row_number,
                    'errors': errors,
                    'data': make_json_serializable(normalized_data)
                }
                if self.import_job.error_details is None:
                    self.import_job.error_details = []
                self.import_job.error_details.append(error_entry)
                self.import_job.add_progress_log(
                    f'Fila {row_number}: Errores de validación - {", ".join(errors)}',
                    'error'
                )
            else:
                try:
                    result = importer_instance.import_action(validated_data)

                    # Handle different return formats from import_action
                    if result is True or result is None:
                        # Legacy format: True/None means created
                        self.import_job.success_rows += 1
                        self.import_job.created_rows += 1
                        if idx % 10 == 0 or idx == len(rows):
                            self.import_job.add_progress_log(
                                f'Procesadas {idx} de {len(rows)} filas...',
                                'info'
                            )
                    elif isinstance(result, str) and result in ['created', 'updated', 'skipped']:
                        # String format: 'created', 'updated', 'skipped'
                        if result == 'skipped':
                            # Skipped rows don't count as success or error
                            pass
                        else:
                            self.import_job.success_rows += 1
                            if result == 'created':
                                self.import_job.created_rows += 1
                            else:  # updated
                                self.import_job.updated_rows += 1

                        if idx % 10 == 0 or idx == len(rows):
                            self.import_job.add_progress_log(
                                f'Procesadas {idx} de {len(rows)} filas ({self.import_job.created_rows} creadas, {self.import_job.updated_rows} actualizadas)...',
                                'info'
                            )
                    elif isinstance(result, dict) and result.get('action') in ['created', 'updated', 'skipped']:
                        # Dict format: {'action': 'created/updated/skipped'}
                        if result['action'] == 'skipped':
                            # Skipped rows don't count as success or error
                            pass
                        else:
                            self.import_job.success_rows += 1
                            if result['action'] == 'created':
                                self.import_job.created_rows += 1
                            else:  # updated
                                self.import_job.updated_rows += 1

                        if idx % 10 == 0 or idx == len(rows):
                            self.import_job.add_progress_log(
                                f'Procesadas {idx} de {len(rows)} filas ({self.import_job.created_rows} creadas, {self.import_job.updated_rows} actualizadas)...',
                                'info'
                            )
                    else:
                        # Any other value is treated as an error
                        self.import_job.error_rows += 1
                        error_entry = {
                            'row': row_number,
                            'errors': [str(result)],
                            'data': make_json_serializable(normalized_data)
                        }
                        if self.import_job.error_details is None:
                            self.import_job.error_details = []
                        self.import_job.error_details.append(error_entry)
                        self.import_job.add_progress_log(
                            f'Fila {row_number}: Error en import_action - {result}',
                            'error'
                        )

                except Exception as e:
                    self.import_job.error_rows += 1
                    error_entry = {
                        'row': row_number,
                        'errors': [f'Excepción: {str(e)}'],
                        'data': make_json_serializable(normalized_data)
                    }
                    if self.import_job.error_details is None:
                        self.import_job.error_details = []
                    self.import_job.error_details.append(error_entry)
                    self.import_job.add_progress_log(
                        f'Fila {row_number}: Excepción - {str(e)}',
                        'error'
                    )

            self.import_job.processed_rows += 1
            self.import_job.save(update_fields=['processed_rows', 'success_rows', 'created_rows', 'updated_rows', 'error_rows', 'error_details'])
