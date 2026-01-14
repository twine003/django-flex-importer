"""
Base class for FlexImporter
"""
from django.db import models
from django.core.exceptions import ValidationError
import csv
import json
from io import StringIO, BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, date
from decimal import Decimal


class FlexImporterMeta:
    """Meta options for FlexImporter"""
    verbose_name = None
    can_re_run = False
    key_field = None


class FlexImporterBase(type):
    """Metaclass for FlexImporter to handle registration"""

    def __new__(mcs, name, bases, attrs):
        cls = super().__new__(mcs, name, bases, attrs)

        if name != 'FlexImporter' and not attrs.get('_abstract', False):
            from .registry import importer_registry
            importer_registry.register(cls)

        return cls


class FlexImporter(metaclass=FlexImporterBase):
    """
    Base class for creating flexible importers.

    Example:
        class SalesImporter(FlexImporter):
            date = models.DateTimeField(verbose_name='Fecha')
            cliente = models.TextField(verbose_name='Cliente')
            producto = models.IntegerField(verbose_name='Producto')

            class Meta:
                verbose_name = "Importador de Ventas"
                can_re_run = True

            def import_action(self, row_data):
                # Custom import logic here
                pass
    """

    _abstract = True

    class Meta:
        verbose_name = None
        can_re_run = False
        key_field = None

    def __init__(self):
        self._setup_meta()

    def _setup_meta(self):
        """Setup meta options"""
        if hasattr(self, 'Meta'):
            self.meta = self.Meta()
        else:
            self.meta = FlexImporterMeta()

    @classmethod
    def get_fields(cls):
        """Get all field definitions from the class"""
        fields = {}
        for name, value in cls.__dict__.items():
            if isinstance(value, models.Field):
                fields[name] = value
        return fields

    @classmethod
    def get_verbose_name(cls):
        """Get the verbose name for the importer"""
        if hasattr(cls, 'Meta') and hasattr(cls.Meta, 'verbose_name'):
            return cls.Meta.verbose_name
        return cls.__name__

    @classmethod
    def can_re_run(cls):
        """Check if the importer can be re-run"""
        if hasattr(cls, 'Meta') and hasattr(cls.Meta, 'can_re_run'):
            return cls.Meta.can_re_run
        return False

    @classmethod
    def get_key_field(cls):
        """Get the key field for update/create operations"""
        if hasattr(cls, 'Meta') and hasattr(cls.Meta, 'key_field'):
            return cls.Meta.key_field
        return None

    @classmethod
    def get_field_info(cls):
        """Get field information for template generation"""
        fields = cls.get_fields()
        field_info = []

        for field_name, field in fields.items():
            verbose_name = getattr(field, 'verbose_name', field_name)
            required = not field.blank if hasattr(field, 'blank') else True
            field_type = cls._get_field_type_name(field)

            field_info.append({
                'name': field_name,
                'verbose_name': verbose_name,
                'required': required,
                'type': field_type,
                'field': field
            })

        return field_info

    @classmethod
    def _get_field_type_name(cls, field):
        """Get a readable field type name"""
        type_map = {
            models.CharField: 'text',
            models.TextField: 'text',
            models.IntegerField: 'integer',
            models.FloatField: 'float',
            models.DecimalField: 'decimal',
            models.BooleanField: 'boolean',
            models.DateField: 'date',
            models.DateTimeField: 'datetime',
            models.EmailField: 'email',
        }

        for field_class, type_name in type_map.items():
            if isinstance(field, field_class):
                return type_name

        return 'text'

    @classmethod
    def generate_template_xlsx(cls):
        """Generate Excel template"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Template"

        field_info = cls.get_field_info()

        for idx, info in enumerate(field_info, start=1):
            header = info['verbose_name']
            if info['required']:
                header += ' *'
            ws.cell(row=1, column=idx, value=header)

        ws2 = wb.create_sheet("Información")
        ws2.cell(row=1, column=1, value="Campo")
        ws2.cell(row=1, column=2, value="Tipo")
        ws2.cell(row=1, column=3, value="Requerido")

        for idx, info in enumerate(field_info, start=2):
            ws2.cell(row=idx, column=1, value=info['verbose_name'])
            ws2.cell(row=idx, column=2, value=info['type'])
            ws2.cell(row=idx, column=3, value='Sí' if info['required'] else 'No')

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @classmethod
    def generate_template_csv(cls):
        """Generate CSV template"""
        field_info = cls.get_field_info()

        output = StringIO()
        writer = csv.writer(output)

        headers = []
        for info in field_info:
            header = info['verbose_name']
            if info['required']:
                header += ' *'
            headers.append(header)

        writer.writerow(headers)

        buffer = BytesIO()
        buffer.write(output.getvalue().encode('utf-8-sig'))
        buffer.seek(0)
        return buffer

    @classmethod
    def generate_template_json(cls):
        """Generate JSON template"""
        field_info = cls.get_field_info()

        template = {
            'template_info': {
                'importer': cls.get_verbose_name(),
                'fields': []
            },
            'data': []
        }

        example_row = {}
        for info in field_info:
            field_def = {
                'name': info['name'],
                'verbose_name': info['verbose_name'],
                'type': info['type'],
                'required': info['required']
            }
            template['template_info']['fields'].append(field_def)
            example_row[info['name']] = f"ejemplo_{info['type']}"

        template['data'].append(example_row)

        buffer = BytesIO()
        buffer.write(json.dumps(template, indent=2, ensure_ascii=False).encode('utf-8'))
        buffer.seek(0)
        return buffer

    @classmethod
    def validate_row(cls, row_data):
        """Validate a single row of data"""
        errors = []
        validated_data = {}
        field_info = cls.get_field_info()

        for info in field_info:
            field_name = info['name']
            field = info['field']
            value = row_data.get(field_name) or row_data.get(info['verbose_name'])

            if info['required'] and (value is None or value == ''):
                errors.append(f"El campo '{info['verbose_name']}' es requerido")
                continue

            if value is not None and value != '':
                try:
                    validated_value = cls._convert_field_value(field, value, info['type'])
                    validated_data[field_name] = validated_value
                except (ValueError, ValidationError) as e:
                    errors.append(f"Error en campo '{info['verbose_name']}': {str(e)}")
            else:
                validated_data[field_name] = None

        return validated_data, errors

    @classmethod
    def _convert_field_value(cls, field, value, field_type):
        """Convert and validate field value"""
        if value is None or value == '':
            return None

        try:
            if field_type == 'integer':
                return int(value)
            elif field_type == 'float':
                return float(value)
            elif field_type == 'decimal':
                return Decimal(str(value))
            elif field_type == 'boolean':
                if isinstance(value, bool):
                    return value
                return str(value).lower() in ('true', 'yes', 'si', 'sí', '1', 't')
            elif field_type == 'date':
                if isinstance(value, date):
                    return value
                return datetime.strptime(str(value), '%Y-%m-%d').date()
            elif field_type == 'datetime':
                if isinstance(value, datetime):
                    return value
                return datetime.fromisoformat(str(value))
            else:
                return str(value)
        except Exception as e:
            raise ValidationError(f"No se pudo convertir el valor '{value}' al tipo {field_type}")

    def import_action(self, row_data):
        """
        Override this method to implement custom import logic.

        Args:
            row_data (dict): Validated row data

        Returns:
            bool or str: True if success, error message if failed
        """
        raise NotImplementedError("Debe implementar el método import_action")
